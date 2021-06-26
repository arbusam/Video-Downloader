import argparse
import youtube_dl
from colorama import init, Fore, Back, Style
from common import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook, load_workbook

def validate_url(url):
    extractors = youtube_dl.extractor.gen_extractors()
    for e in extractors:
        if e.suitable(url) and e.IE_NAME != "generic":
            return True

def display_arguments():
    parser = argparse.ArgumentParser(
        description=f"This application {Fore.LIGHTMAGENTA_EX}stores{Fore.RESET} one or more YouTube URL(s) into the {Fore.LIGHTMAGENTA_EX}database{Fore.RESET}.\nYou must add the links into the database {Fore.LIGHTMAGENTA_EX}before{Fore.RESET} they can be downloaded.",
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # 'url' argument
    parser.add_argument("-u", "--url", type=str, help="YouTube URL (link)")

    # 'file' argument
    parser.add_argument(
        "-f",
        "--file",
        type=str,
        help="file containing YouTube URLs (one URL per line)",
    )

    # 'type' argument
    parser.add_argument(
        "-t",
        "--type",
        type=str,
        default=default_download_type,
        help=f"""
        Download type.
        Options are {get_type_options()}""",
    )
    return parser

def get_type_options():
    options = ""
    for i in range(1, len(file_types)+1):
        options + file_types.keys()[i]

def validate_download_types(args):
    download_types = []
    invalid = False
    for download_type in args.type.split("+"):
        if download_type == "all":
            download_types = file_types
            break
        if download_type in file_types:
            download_types.append(download_type)
        else:
            print(INVALID_DOWNLOAD_TYPE.format(type=download_type))
            invalid = True
            download_types = default_download_type.split("+")
            break
    return [download_types, invalid]


# =============================================
# Return a list containing all the file entries
# =============================================
def get_links_from_file(file_name):
    # create a blank list to add all validated links
    file_entries = []
    try:
        f = open(file_name, "r")
    except FileNotFoundError:
        # terminate with an error message
        exit(FILE_DOES_NOT_EXIST.format(file=file_name))
    else:
        file_entries = f.readlines()
        f.close()
        print(FOUND_ENTRIES_IN_FILE.format(num=len(file_entries), file=file_name))
        # return the list of all URLs
        return file_entries


# ======================================
# Validate all URLs using youtube_dl API
# ======================================
def validate_links(links):
    # create a blank list to add all validated links
    valid_urls = []
    # iterate through the links
    for link in links:
        # let's clean-up each string and remove any spaces / paddings
        link = link.strip()

        extractors = youtube_dl.extractor.gen_extractors()
        # loop through all extractors
        for e in extractors:
            # if link is 'suitable', then add it to 'valid_urls' list
            if e.suitable(link) and e.IE_NAME != "generic":
                valid_urls.append(link)
                print(LINK_VALID.format(link=link))
                # break out of the extractor loop and validate the next 'link'
                break
        else:
            # 'else' gets executed AFTER the 'for' loop (unless you 'break' out of the for loop)
            print(LINK_INVALID.format(link=link))

    # only keep unique values (remove repeats)
    unique_valid_urls = []
    # a 'set' is an UNORDERED collections of UNIQUE elements
    unique_valid_urls = list(set(valid_urls))

    # finally, return the list containing the valid urls
    return unique_valid_urls 

# ======================================
# Excel
# ======================================
def populate_headings(ws, values):
    """Populates the heading for a given worksheet

    Args:
        ws (Worksheet): Worksheet object
        values (List): List of dictionaries containing the headings

    Returns:
        Worksheet: Worksheet object with headings populated
    """

    print(Fore.YELLOW + "Populating Excel headings...", end="\r")

    for i in range(len(values)):
        ws[get_column_letter(i+1) + "1"].value = values[i]["name"]
        ws[get_column_letter(i+1) + "1"].alignment = Alignment(
            horizontal=values[i]["h_align"],
            vertical=values[i]["v_align"]
        )
        ws[get_column_letter(i+1) + "1"].fill = PatternFill(
            fgColor=values[i]["bg_color"], fill_type="solid"
        )
        ws[get_column_letter(i+1) + "1"].font = Font(
            name=values[i]["font_name"],
            size=values[i]["font_size"],
            bold=values[i]["bold"],
            color=values[i]["text_color"],
        )
        
        ws.column_dimensions[get_column_letter(i+1)].width = values[i]["column_size"]
    print(Fore.GREEN + "Excel headings populated")
    return ws

def populate_data(ws, data):
    """Populates the data for a given worksheet

    Args:
        ws (Worksheet): Worksheet object
        data (List): List of lists which each contain the data for one row

    Returns:
        Worksheet: Worksheet object with headings populated
    """

    for row in range(len(data)):
        for column in range(len(data[row])):
            ws[get_column_letter(column+1) + str(row+2)].value = data[row][column]